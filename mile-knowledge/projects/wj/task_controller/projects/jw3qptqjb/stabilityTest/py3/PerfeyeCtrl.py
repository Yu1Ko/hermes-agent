# -*- coding: utf-8 -*-

import random
import time

import openpyxl
import pandas
import zipfile
from testplus import *
from BaseToolFunc import *


PERF_DIR = r'LOG'
PERFEYE_VER = 'Perfeye-2.1.11-release'

x = os.path.realpath(__file__)
root = x.split('\\')[0]
PERFMON_PATH_LOCAL = os.path.join(root, os.sep, 'PerfMon3')
logger = logging.getLogger(str(os.getpid()))

def un_zip(file_name, dst):
    """解压 zip 文件"""
    zip_file = zipfile.ZipFile(file_name)
    if os.path.isdir(dst):
        pass
    else:
        os.makedirs(dst)
    for names in zip_file.namelist():
        zip_file.extract(names, dst)
    zip_file.close()


class PerfeyeData(object):
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.checkFolder(folder_path)
        self.dicPerfeyeData = {}
        self.dicPerfmonData = {}
        self.readPerfeyeXlsx()

    def checkFolder(self, folder_path):
        if not os.path.exists(folder_path):
            raise Exception('no folder_path:{}'.format(folder_path))
        report_file_path = os.path.join(folder_path, 'report.xlsx')
        if not os.path.exists(report_file_path):
            raise Exception('no report.xlsx in {}'.format(folder_path))
        self.report_file_path = report_file_path
        zip_file_path = os.path.join(folder_path, 'report.capture.zip')
        if not os.path.exists(zip_file_path):
            raise Exception('no report.capture.zip in {}'.format(folder_path))
        self.zip_file_path = zip_file_path


    def readPerfeyeXlsx(self):
        file_name = self.report_file_path
        wb_obj = openpyxl.load_workbook(file_name)
        sheet = wb_obj.active
        # print(sheet.max_row, sheet.max_column)
        for y in range(sheet.max_column):
            y = y+1
            if  sheet.cell(4, y).value is None:
                break
            # self.dicDeviceInfo[sheet.cell(4,y).value] = sheet.cell(5,y).value

        for y in range(sheet.max_column):
            y = y+1
            if  sheet.cell(12, y).value is None:
                break
            self.dicPerfeyeData[sheet.cell(12, y).value] = []
            for x in range(13, sheet.max_row + 1):
                if sheet.cell(x, 1).value is None:
                    break
                self.dicPerfeyeData[sheet.cell(12,y).value].append(sheet.cell(x, y).value)

    def trans2PerfmonData(self):
        listFilterKeys = ['FTime', 'Num','time(ms)', 'label','Notes','Jank','BigJank']
        dicTransKeys = { 'absTime':'Time', 'FPS':'RLFPS', 'Memory(MB)':'虚拟内存(KB)','Total(%)':'系统CPU占用率(%)',
                         'App(%)':'CPU占用率(%)','GPULoad':'GPU Load(%)'}
        for key in self.dicPerfeyeData:
            if key in listFilterKeys:
                continue
            if 'FTime' in key:
                continue
            if key in dicTransKeys:
                self.dicPerfmonData[dicTransKeys[key]] = self.dicPerfeyeData[key]
                if key == 'Memory(MB)':
                    new_list = []
                    for v in self.dicPerfmonData[dicTransKeys[key]]:
                        new_list.append(v*1024)
                    self.dicPerfmonData[dicTransKeys[key]] = new_list
            else:
                self.dicPerfmonData[key] = self.dicPerfeyeData[key]
        # process FrameTime
        self.dicPerfmonData['FrameTime'] = []
        for i in range(len(self.dicPerfeyeData['FTime1(ms)'])):
            dd = self.dicPerfeyeData['FTime1(ms)'][i]
            if dd is None:
                strFrameTime = '0'
            else:
                strFrameTime = str(int(self.dicPerfeyeData['FTime1(ms)'][i]))
            for j in range(2,100):
                tmp_key = 'FTime{}(ms)'.format(j)
                if tmp_key not in self.dicPerfeyeData:
                    break
                if self.dicPerfeyeData[tmp_key][i] is None:
                    break
                strFrameTime = strFrameTime + ','+str(int(self.dicPerfeyeData[tmp_key][i]))
            self.dicPerfmonData['FrameTime'].append(strFrameTime)

        # Time格式转换
        listTime = []
        for absTime in self.dicPerfmonData['Time']:
            timeArray = time.localtime(absTime / 1000)
            datetime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
            listTime.append(datetime)
        self.dicPerfmonData['Time'] = listTime


    def save_perfmon_data(self):
        self.trans2PerfmonData()
        df_perfmon = pandas.DataFrame(self.dicPerfmonData, dtype=object)
        summaryPath = os.path.join(self.folder_path, 'sys_summary.tab')
        df_perfmon.to_csv(summaryPath, sep='\t', index=None, encoding="utf_8")
        self.summaryPath = summaryPath

    def save_perfmon_screenshot(self):
        #zip screenshot
        zip_file = self.zip_file_path
        un_zip(zip_file, self.folder_path)

    def del_perfeye_files(self):
        filecontrol_deleteFileOrFolder(self.report_file_path)
        filecontrol_deleteFileOrFolder(self.zip_file_path)

class PerfeyeControl(object):
    def __init__(self, deviceId, bPerfeyeTest=False, strMachineTag='Test',nScreenshot_Interval=2,strPackageName=None,strAppKey='yizhiban'):
        self.initLogger()
        self.deviceId=deviceId
        self.bPerfeyeTest=bPerfeyeTest
        self.strMachineTag=strMachineTag
        self.strPackageName=strPackageName
        self.testplus=None
        self.nScreenshot_Interval=nScreenshot_Interval
        self.bMobile=True
        self.strAppKey=strAppKey
        self.bPerfeyeStop=False
        self.bPerfeyeSave=False
        if self.strMachineTag=='PC':
            self.bMobile=False

    def initLogger(self):
        try:
            initLog(self.__class__.__name__)
            self.log = logging.getLogger(str(os.getpid()))
        except Exception:
            info = traceback.format_exc()
            # print(info)
            raise Exception('initLogger ERROR!!')

    def PerfeyeCreate(self):
        # 临时Perfey测试环境
        if self.bPerfeyeTest:
            Perfeye_ver = "Perfeye-2.1.3-release-test"
        elif self.strMachineTag == 'Test':
            Perfeye_ver = PERFEYE_VER + '-' + self.deviceId
        else:
            Perfeye_ver = PERFEYE_VER
        # Step 1. 打开TPT客户端
        if sys.platform.startswith('win'):
            x = os.path.realpath(__file__)
            root = x.split('\\')[0]
            perfeye_exe_path = os.path.join(root, os.sep, Perfeye_ver, 'Perfeye.exe')
            self.testplus = Testplus(perfeye_path=perfeye_exe_path)
        else:
            perfeye_exe_path = filecontrol_find_folder_path('/', 'perfeye-linux')[0]
            self.testplus = Testplus(perfeye_path=perfeye_exe_path,
                                python=os.path.join(perfeye_exe_path, 'venv', 'bin', 'python3'))
        self.log.info(perfeye_exe_path)
        result, error = self.testplus.launch(port=random.randint(10000, 60000), headless=True, timeout=100)
        if not result:
            self.testplus.kill()
            raise Exception('perfeye CreateTPTClient error:{}'.format(error))

    def PerfeyeConnect(self):
        # Step 4. 连接手机
        # PC端package_name必须传None
        dic_configs={}
        if not self.bMobile:
            #PC端采集 Vulkan引擎数据
            dic_configs={"old_jank_counter": True,"graphics_api": 0}
        result, error = self.testplus.connect(self.deviceId, self.strPackageName,configs=dic_configs)
        self.log.info('perfeye connect:{}'.format(result))
        if not result:
            self.testplus.kill()
            raise Exception('perfeye Connect error:{}'.format(error))

        #self.PerfeyeCheckDevice()  2.1.1默认采集CPU温度

    def PerfeyeCheckDevice(self):
        #目前只有android需要调用该接口用于采集cpu温度  需要在Perfeyestart前调用
        if self.strMachineTag=='Android':
            self.testplus.android_check_device(self.deviceId)

    def PerfeyeStart(self,nPid=None):
        '''
        Android
```json
  {
      CPU_USAGE               : 1,  // cpu使用率
      CORE_FREQUENCY          : 2,  // cpu频率
      GPU_USAGE               : 3,  // gpu使用率
      GPU_FREQ                : 4,  // gpu频率
      FPS                     : 5,  // 帧率(必须项)
      NETWORK_USAGE           : 6,  // 网络
      SCREEN_SHOT             : 8,  // 截图
      MEMORY                  : 9,  // 内存
      BATTERY                 : 10, // 电池数据
      CPU_TEMPERATURE         : 11, // cpu温度
      FRAME_TIME              : 12, // 帧间隔(必须项)
      ANDROID_MEMORY_DETAIL   : 13, // 内存详情
      CORE_USAGE              : 14, // cpu多核使用率
      BATTERY_TEMPERATURE     : 19  // 电池温度
      GPU_TEMPERATURE         : 35, // gpu温度
  }
```
iOS
```json
  {
      CPU_USAGE               : 1,  // cpu使用率
      FPS                     : 5,  // 帧率(必须项)
      NETWORK_USAGE           : 6,  // 网络
      SCREEN_SHOT             : 8,  // 截图
      MEMORY                  : 9,  // 内存
      FRAME_TIME              : 12, // 帧间隔(必须项)
      IOS_GPU_USAGE           : 17, // IOS gpu使用率
      CTX_SWITCH              : 15, // IOS SWITCH
      WAKEUP                  : 16, // IOS 唤醒次数
      IOS_ENERGY_USAGE        : 18, // IOS 能耗
      BATTERY_TEMPERATURE     : 19, // IOS 电池温度
      IOS_IO                  : 10001, //IOS  IO读写使用
  }
```

win
```json
  {
      CPU_USAGE               : 1,  // cpu使用率
      CORE_FREQUENCY          : 2,
      GPU_USAGE               : 3,
      GPU_FREQ                : 4,  // gpu频率
      FPS                     : 5,  // 帧率(必须项)
      NETWORK_USAGE           : 6,  // 网络
      SCREEN_SHOT             : 8,  // 截图
      MEMORY                  : 9,  // 内存
      CPU_TEMPERATURE         : 11, // cpu温度
      FRAME_TIME              : 12, // 帧间隔(必须项)
      CORE_USAGE              : 14,
      RENDERER                : 32,
      IO                      : 33,
      LOGIC_FPS               : 34,
      GPU_TEMPERATURE         : 35,
      BOARD_POWER_DRAW        : 36,
      DRAM_BANDWIDTH          : 37, // DRAM_BANDWIDTH
      CPU_CORE_IPC            : 38, // CPU_Core_IPC
      CPU_CORE_FREQ           : 39, // CPU_Core_Freq
      CPU_CORE_CACHE_MISS     : 40, // CPU_CORE_CACHE_MISS
      CPU_CORE_CACHE_HIT_RATIO: 41, // CPU_CORE_CACHE_HIT_RATIO
      CPU_CORE_CACHE_MPI      : 42, // CPU_CORE_CACHE_MPI
      CPU_CORE_TEMP           : 43, // CPU_Core_Temp
      CPU_ENERGY              : 44, // CPU_Energy
      GPA_RENDERER            : 46 // GPA_RENDERER
  }
        '''
        # Step 5. 开始测试
        if self.nScreenshot_Interval:
            nScreenshot_Interval=self.nScreenshot_Interval
        else:
            nScreenshot_Interval = 2
        if self.strMachineTag=='Android':
            data_types = [1, 2, 3, 4, 5, 6, 8, 9,10, 11, 12, 13, 14,19,35]
            # ios端30采集截图
        elif self.strMachineTag == 'Ios':
            #17,18,19,10001
            if not self.nScreenshot_Interval:
                nScreenshot_Interval = 15
            data_types = [1,5, 8, 9, 12,17,19]
        else:
            data_types = [1, 2, 3, 4, 5, 8, 9, 11, 12, 14, 32, 34, 35, 36,47]

        if self.bMobile:
            result, error = self.testplus.start(self.deviceId, package_name=self.strPackageName,data_types=data_types ,screenshot_interval=nScreenshot_Interval)
        else:
            result, error = self.testplus.start(self.deviceId, pid=nPid,data_types=data_types,screenshot_interval=nScreenshot_Interval)

        self.log.info('perfeye start:{}'.format(result))
        if not result:
            self.testplus.kill()
            raise Exception('perfeye Start error:{}'.format(error))

    def PerfeyeSetTimeNode(self):
        result, error = self.testplus.start_test_case(self.deviceId,'')
        self.log.info('perfeye start_test_case:{}'.format(result))
        if not result:
            raise Exception('perfeye start_test_case error:{}'.format(error))

    def PerfeyeStop(self):
        # Step 7. 停止测试
        if self.bPerfeyeStop:
            return
        self.bPerfeyeStop=True
        result, error = self.testplus.stop(self.deviceId)
        self.log.info('perfeye stop:{}'.format(result))
        if not result:
            raise Exception('perfeye Stop error:{}'.format(error))

    def PerfeyeKill(self):
        # 结束进程
        if self.testplus:
            self.testplus.kill()
            # 结束adb设备端口
            if '-' not in self.deviceId and self.deviceId != 'localhost':
                adb_forward_remove(self.deviceId)

    def PerfeyeSave(self,subtags, extraData={}, BVT=True,nMaxRetransmissionCount=3,strVersion='1.0.0'):
        if self.bPerfeyeSave:
            return
        #未Save过  先判断一下Stop
        self.PerfeyeStop()
        self.bPerfeyeSave=True
        if extraData and not extraData["datalist"]:
            extraData = {}
        if self.strAppKey == 'jx3hd':
            self.strAppKey = 'JX3'
        elif self.strAppKey == 'jx3classic':
            self.strAppKey = 'JX3CLASSIC'
        # Step 8. 保存数据
        folder_datetime = time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime(time.time()))
        if sys.platform.startswith('win'):
            root = os.path.realpath(__file__).split(STRSEPARATOR)[0]
            folder_path = os.path.join(root, os.sep, 'PerfeyeDataSave', folder_datetime)
        else:
            folder_path = os.path.join('/home/test', 'PerfeyeDataSave', folder_datetime)

        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        report_file_path = os.path.join(folder_path, 'report.xlsx')
        listTags = subtags.split('|')  # subtags such as:10|PC|RTX2080|藏剑山庄_乱世|campfight|2022-08-16
        myScenes = '{}|{}'.format(listTags[3], listTags[4])
        self.log.info("saveStart:{}".format(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
        #strVersion=get_package_version()
        result=False
        iRetransmissionCount = 0
        #nMaxRetransmissionCount = 3
        nRetSaveData=''
        nRetUploadData=''
        error=''
        while True:
            if not result:
                ''''''
                nRetSaveData = error.find('保存数据成功')
                nRetUploadData = error.find('上传数据成功')
                if nRetUploadData != -1:
                    return True
                else:
                    if iRetransmissionCount >= nMaxRetransmissionCount:
                        logger.info('perfeye Save error:{}'.format(error))
                        return False
                    if nMaxRetransmissionCount!=1 and iRetransmissionCount == nMaxRetransmissionCount-1:
                        extraData={}
                    iRetransmissionCount += 1
                    if iRetransmissionCount>1:
                        self.log.info('重传次数:{}'.format(iRetransmissionCount))
                    time.sleep(30 * iRetransmissionCount)
                    strTime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
                    logger.info("saveStart retry{}:{}".format(iRetransmissionCount, strTime))
                    result, error = self.testplus.save(self.deviceId, case_name=subtags, scenes=myScenes,  # picture_quality='画质',
                                                  report_file_path=report_file_path,
                                                  need_save=False, need_upload=True, appKey=self.strAppKey, do_upload=BVT,
                                                  extra_data=extraData,version=strVersion)
                    self.log.info("saveEnd retry{}:{}".format(iRetransmissionCount,
                                                            time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
            else:
                break
        self.log.info("saveEnd:{}".format(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
        # processDataForPerfmon(folder_path)
        self.PerfeyeKill()
        return True
        pass

if __name__ == '__main__':
    pass

