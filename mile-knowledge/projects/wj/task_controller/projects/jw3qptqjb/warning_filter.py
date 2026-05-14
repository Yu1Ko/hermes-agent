import os
import re
import csv
import pandas as pd
from BaseToolFunc import *
# 输出日期
from datetime import datetime
source="E:\AUTO_BVT_NEW\case\liuzhu\py3\TempFolder\warning\\"
now = datetime.now()
formatted_date = now.strftime("%Y-%m-%d")
if not filecontrol_existFileOrFolder(source+formatted_date):
    filecontrol_createFolder(source+formatted_date)

source+=formatted_date+'\\'
#最终文件
strFileName = source+f"{formatted_date}_warning.xlsx"


source1="E:\AUTO_BVT_NEW\case\liuzhu\py3\TempFolder\warning\\"

#input_file = source1+f"{now.month}-{now.day}_warning.txt"
input_file=SERVER_PATH + f"\Warning\{formatted_date}"+r"\full-log.txt"

output_file = source+"warnings.csv"
output_file_include_dir = source+"warnings_include_dir.csv"
output_file_exclude = source+"warnings_exclude.csv"


output_sorted = source+'warning_sorted_filename.csv'
output_sorted_include_dir = source+'warning_sorted_filename_include_dir.csv'
output_sorted_exclude = source+'warning_sorted_filename_exclude.csv'
def test():

    included_dirs = ['\\Sword3\\Source\\KG3DEngineDX11',
                     '\\Sword3\\Source\\KG3DEngine\\',
                     '\\Sword3\\Source\\Common\\Schema\\Semantic']  #包含的目录
    for dir in included_dirs:
        print(dir)
    excluded_dirs = [r'\Sword3\Source\KG3DEngineDX11\KG3DEngineE\Internal\Tool',
                     r'\Sword3\Source\KG3DEngineDX11\KG3DEngineE\Internal\Module\KG3DEngineDX11E\EditTool',
                     r'\Sword3\Source\KG3DEngineDX11\KG3DEngineE\Internal\Module\IPPVideo\detours',
                     r'\Sword3\Source\KG3DEngineDX11\DevEnv',
                     r'\Sword3\Source\KG3DEngineDX11\KG3DEngineE\Internal\Component\KG3D_DebugConsole',
                     r'\Sword3\Source\KG3DEngine\KG3DEngine\JpegLib',
                     r'\Sword3\Source\KG3DEngineDX11\KG3DEngineE\Internal\Module\KG3D_NVBlastDestruction\MeshFix',
                     r'\Sword3\Source\KG3DEngineDX11\KG3DEngineE\Internal\Module\IPPVideo\ffmpeg',
                     r'\Sword3\Source\KG3DEngine\KG3DSound\KMp3LibClass\src\KMp3HufTable.h',
                     r'\Sword3\Source\KG3DEngineDX11\KG3DEngineE\Internal\Module\KG3D_Wwise',
                     r'\Sword3\Source\KG3DEngineDX11\KG3DEngineE\Publish\Include\IKG3D_WwiseInterface.h',
                     r'\Sword3\Source\KG3DEngine\KG3DEngineCommon\kg3d_DataCenter\datacenter\external\boost',
                     r'\Sword3\Source\KG3DEngine\KG3DEngineCommon\kg3d_DataCenter\datacenter\lua',
                     r'\Sword3\Source\KG3DEngine\KG3DEngineCommon\include\KG3D_DataCenter\datacenterpub\external'] #去掉的目录



    excluded_Cwarning = ['C4189', 'C26475', 'C26814','C26440','C4100','C4251']
    excluded_StrWarning=["(con.4)",
                         "Use 'nullptr' rather than 0 or NULL (es.47)",
                         "Avoid 'goto' (es.76)", "Avoid 'goto' (es.76)",
                         "Don't use C-style casts (type.4)",
                         "Consider using gsl::finally if final action is intended (gsl.util)",
                         "Default constructor may not throw. Declare it 'noexcept' (f.6)",
                         "Prefer to use gsl::at() instead of unchecked subscript operator (bounds.4)"
                         ]

    header = ['id', 'file', 'row', 'col','warning']    # 定义表格的列名
    with open(output_file, 'w', newline='') as f: # 创建空的 CSV 文件并写入表头
        writer = csv.writer(f)
        writer.writerow(header)

    with open(output_file_include_dir, 'w', newline='') as f: # 创建空的 CSV 文件并写入表头
        writer = csv.writer(f)
        writer.writerow(header)

    with open(output_file_exclude, 'w', newline='') as f: # 创建空的 CSV 文件并写入表头
        writer = csv.writer(f)
        writer.writerow(header)


    data_hold_exclude = []
    data_hold_include_dir = []
    data_hold=[]
    with open(input_file, 'r') as file:
        for line in file:
            if 'warning' in line:
                data={}
                match = re.search(r'warning\s+(C\d+):', line)  #找警告类型
                if match:
                    warning_type = match.group(1)
                    data['id'] = warning_type
                    '''
                    if any(dir in warning_type for dir in excluded_Cwarning): #过滤掉C类警告
                        continue
                    '''
                else:
                    continue

                match = re.search(r'[A-Za-z]:\\[\\\w\s\-.]*\.\w+', line) # 找路径
                if match:
                    path = match.group()
                    match_ = re.search(r'\\Sword3\\(.*)', path)  # 匹配从sword3开始到最后
                    if match_:
                        path = match_.group()
                        data['file'] = path
                    else:
                        continue
                else:
                    match1 = re.search(r'[A-Za-z]:\\(?:[^\\/:*?"<>|\r\n]+\\)*([^\\/:*?"<>|\r\n]+)', line) #匹配没有后缀的路径
                    if match1:
                        path = match1.group()
                        match2 = re.search(r'\\Sword3\\(.*)', path)  # 匹配从sword3开始到最后
                        if match2:
                            path = match2.group()
                            data['file'] = path
                        else:
                            continue
                    else:
                        data['file'] = " "
                        print("有空白路径")

                match = re.search(r'\((\d+),(\d+)\):', line)   #行和列
                if match:
                    row = str(match.group(1))
                    col = str(match.group(2))
                    data['row'] = row
                    data['col'] = col
                else:
                    match_ = re.search(r'\((\d+)\):', line)
                    if match_:
                        row = match_.group(1)
                        data['row'] = row
                        data['col'] = " "
                    else:
                        data['row'] = " "
                        data['col'] = " "
                        print("有空白行列")

                match = re.search(r'warning\s+([^:]+):\s+(.*)', line)  #警告内容
                if match:
                    warning_info = match.group(2)
                    '''
                    if any(str in warning_info for str in excluded_StrWarning):  # 过滤掉excluded_StrWarning里面的警告
                        continue
                    '''
                    data['warning'] = warning_info
                else:
                    data['warning'] = " "
                    print("有空白警告")

                '''
                if any(dir in data_exclude['file'] for dir in included_dirs) and not any(   #这里是有些bug吗？？？？
                        dir in data_exclude['file'] for dir in excluded_dirs):
'''

                data_hold.append(data)
                if any(dir in data['file'] for dir in included_dirs): #include dir
                    data_hold_include_dir.append(data)
                    if any(dir in data['file'] for dir in excluded_dirs): #过滤路径
                        continue

                    if any(dir in warning_type for dir in excluded_Cwarning):  # 过滤掉C类警告
                        continue

                    if any(str in warning_info for str in excluded_StrWarning):  # 过滤掉excluded_StrWarning里面的警告
                        continue
                    data_hold_exclude.append(data)


    with open(output_file, 'a', newline='') as f:
        writer = csv.writer(f)
        for row in data_hold:
            writer.writerow([row['id'], row['file'], row['row'], row['col'], row['warning']])  # 写入每一行数据

    with open(output_file_include_dir, 'a', newline='') as f:
        writer = csv.writer(f)
        for row in data_hold_include_dir:
            writer.writerow([row['id'], row['file'], row['row'], row['col'], row['warning']])  # 写入每一行数据

    with open(output_file_exclude, 'a', newline='') as f:
        writer = csv.writer(f)
        for row in data_hold_exclude:
            writer.writerow([row['id'], row['file'], row['row'], row['col'], row['warning']])  # 写入每一行数据

    print(f"data_hold len:{len(data_hold)}")
    print(f"data_hold_include_dir len:{len(data_hold_include_dir)}")
    print(f"data_hold_exclude len:{len(data_hold_exclude)}")


def delete_duplicates_and_sortbyfile():
    df = pd.read_csv(output_file)
    df.drop_duplicates(inplace=True)
    sorted_df = df.sort_values(by='file')
    sorted_df.to_csv(output_sorted, index=False)
    print(f"data_hold_duplicates len:{len(df)}")


    df = pd.read_csv(output_file_include_dir)
    df.drop_duplicates(inplace=True)
    sorted_df = df.sort_values(by='file')
    sorted_df.to_csv(output_sorted_include_dir, index=False)
    print(f"data_hold_include_dir_duplicates len:{len(df)}")

    df = pd.read_csv(output_file_exclude)
    df.drop_duplicates(inplace=True)
    sorted_df = df.sort_values(by='file')
    sorted_df.to_csv(output_sorted_exclude, index=False)
    print(f"data_hold_exclude_duplicates len:{len(df)}")



def get_last_modified_name(strFilePath, nLineNumber,list_lines,bFlag=False):
    try:
        # 运行 svn blame 命令并捕获输出
        if not bFlag:
            strFilePath = strFilePath.replace('\\', '/')
            strFilePath = strFilePath[7:]
            print(strFilePath)
            strURL = 'https://xsjreposvr1.seasungame.com/svn/Sword3/trunk' + strFilePath
            #list_command = ["svn", "blame",'--use-merge-history',strURL]
            command = f'svn blame --use-merge-history {strURL} -x --ignore-eol-style'
            #command = f'svn blame --use-merge-history {strURL} -x --ignore-eol-style'
            pi = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE)
            res = pi.stdout.read()
            try:
                res = str(res, encoding='gbk')
            except:
                res = str(res, encoding='utf8')
            #print(res)
            # 按行分割输出内容
            list_lines = res.splitlines()

        # 获取指定行的信息

        target_line = list_lines[nLineNumber - 1]
        # 提取最后一次修改的人员名称
        list_name=target_line.split(' ')
        '''
        print(list_name)
        for i in range(1,len(list_name)):
            if re.match(r'\d{1,9}',list_name[i]):
                break
        #print(i)
        for j in range(i+1,len(list_name)):
            if list_name[j]!='':
                break
        strName = list_name[j]
        '''
        list_name_new=[]
        for strName in list_name:
            if strName!='':
                list_name_new.append(strName)
        #print(list_name)
        strName=list_name_new[1]
        if list_name_new[0]=='G':
            strName=list_name_new[2]
        print(f"name:{strName}  number:{nLineNumber}  target_line:{target_line}----")
        #print(target_line.split(' '))
        return strName,list_lines
    except (IndexError, subprocess.CalledProcessError) as e:
        print(f"Error: {e}  path: {strFilePath}  number:{nLineNumber}")
        return "代码更新",list_lines

def tt():
    header = ['Id', 'File','Count','Row', 'Col','Lastest modified','Warning']  # 定义表格的列名

    df = pd.read_csv(output_sorted_exclude)
    from openpyxl import Workbook
    from openpyxl.styles import Font
    # 合并单元格
    from openpyxl.styles import Alignment
    # 创建居中格式的样式对象
    center_alignment = Alignment(horizontal='left', vertical='center')
    # 创建加粗的字体样式对象
    bold_font = Font(bold=True)

    wb = Workbook()
    ws = wb.active

    list_fileInfo=[]
    # 将DataFrame数据写入Excel工作表
    for i in range(1,len(header)+1):
        ws.cell(row=1, column=i).value =header[i-1]
        ws.cell(row=1, column=i).font=bold_font
    for r, row in enumerate(df.values, start=1):
        for c, value in enumerate(row, start=1):
            if c>=3:
                c=c+1
            if c>=6:
                c=c+1
            ws.cell(row=r+1, column=c).value = value

    #count列全部置为1
    strPrevFilePath=""
    list_lines=[]
    for i in range(2,ws.max_row + 1):
        ws.cell(row=i, column=3).value = 1
        strFilePath=ws.cell(row=i, column=2).value
        nLineNumber=int(ws.cell(row=i, column=4).value)
        try:
            if strPrevFilePath!=strFilePath:
                strPrevFilePath=strFilePath
                strName,list_lines =get_last_modified_name(strFilePath,nLineNumber,list_lines,False)
            else:
                strName,list_lines =get_last_modified_name(strFilePath,nLineNumber,list_lines,True)

            ws.cell(row=i, column=6).value=strName
        except:
            pass


    fileNameLast = None
    nIndexStart = 0
    nIndexEnd = 0
    nRowCounter=0
    print(ws, ws.max_row + 1)
    #处理最后一行 2
    for i in range(2, ws.max_row + 2):
        if not fileNameLast:
            fileNameLast = ws[f"B{i}"].value
            print(fileNameLast)
            nIndexStart = i
            nIndexEnd = i
            #处理第一行
            list_fileInfo.append({"file": ws[f'B{nIndexStart}'].value, "count":1})
            continue
        fileName = ws[f"B{i}"].value
        if fileName == fileNameLast:
            nIndexEnd += 1
        else:
            ws.merge_cells(f'B{nIndexStart}:B{nIndexEnd}')
            ws.merge_cells(f'C{nIndexStart}:C{nIndexEnd}')
            print(f'B{nIndexStart}:B{nIndexEnd}')
            #单元格设置居中
            ws[f'B{nIndexStart}'].alignment = center_alignment
            ws[f'C{nIndexStart}'].alignment = center_alignment
            ws[f'C{nIndexStart}'].value=nIndexEnd-nIndexStart+1
            list_fileInfo.append({"file":ws[f'B{nIndexStart}'].value,"count":nIndexEnd-nIndexStart+1})
            #重置状态
            fileNameLast = fileName
            nIndexStart = i
            nIndexEnd = i
    # 保存Excel文件

    #设置file列名称长度
    ws.column_dimensions['B'].width = 120
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 150


    #统计页面
    header = ['file', 'count']
    wsCount = wb.create_sheet(title="fileCount")
    for i in range(1,len(header)+1):
        wsCount.cell(row=1, column=i).value =header[i-1]
        wsCount.cell(row=1, column=i).font=bold_font

    #自动定义排序规则
    def custom_sort_rule(dic_info):
        return dic_info['count']

    list_fileInfo.sort(key=custom_sort_rule,reverse=True)
    print(list_fileInfo)
    for i in range(2,len(list_fileInfo)+1):
        for j in range(len(header)):
            print(list_fileInfo[i-2][header[j]])
            wsCount.cell(row=i, column=j+1).value = list_fileInfo[i-2][header[j]]
            wsCount.cell(row=i, column=j+1).alignment= center_alignment

    wsCount.column_dimensions['A'].width = 130
    wb.save(strFileName)


def CopyDataToServer():
    strServerPath = SERVER_PATH + f"\Warning\{formatted_date}"
    if not filecontrol_existFileOrFolder(strServerPath):
        filecontrol_createFolder(strServerPath)
    strLocalPath = "E:\AUTO_BVT_NEW\case\liuzhu\py3\TempFolder\warning\\" + formatted_date
    print(strServerPath)
    filecontrol_copyFileOrFolder(strLocalPath, strServerPath)


def dealWith_data():
    test()
    delete_duplicates_and_sortbyfile()
    # 文件数量
    tt()
    CopyDataToServer()

if __name__ == '__main__':
    dealWith_data()
    '''w
    bDealWithFlag=True 
    while True:
        if bDealWithFlag:
            time_now = datetime.now()
            nHour = time_now.hour
            nMinute = time_now.minute
            if nHour == 9 and nMinute > 30:
                try:
                    dealWith_data()
                    bDealWithFlag=False
                except:
                    sleep_heartbeat(10)
        else:
            sleep_heartbeat(30)

'''



