# coding=utf-8
import csv
import os.path
import time

import pandas as pd
from CaseCommon import *
from BaseToolFunc import *
from datetime import datetime

class CaseEngineWarning(CaseCommon):
    def __init__(self):
        CaseCommon.__init__(self)

    def SetFilePath(self,strFormattedDate):
        # 本地warning文件夹路径
        self.strSource = f"{os.getcwd()}{os.sep}warning{os.sep}"
        # 当前日期
        self.strFormattedDate = strFormattedDate
        # 在本地创建每天存放数据的warning文件夹
        if not filecontrol_existFileOrFolder(self.strSource + self.strFormattedDate):
            filecontrol_createFolder(self.strSource + self.strFormattedDate)
        # 本地每天存放数据的warning文件夹路径
        self.strSource += self.strFormattedDate + os.sep
        # 最终处理后的文件名称
        self.strFileName = self.strSource + f"{self.strFormattedDate}_warning.xlsx"

        # self.strInputSrc = self.strLocalPath+f"{now.month}-{now.day}_warning.txt"
        # 服务端存放日志文件路径
        self.strInputSrc = SERVER_PATH + f"{os.sep}Warning{os.sep}{self.strFormattedDate}{os.sep}full-log.txt"

        # 提取日志中的原始数据
        self.strOutput_file = self.strSource + "warnings.csv"
        # 处理included_dirs后的数据
        self.strOutput_file_include_dir = self.strSource + "warnings_include_dir.csv"
        # 处理excluded后的数据
        self.strOutput_file_exclude = self.strSource + "warnings_exclude.csv"

        # 提取日志中的原始数据（去除重复数据）
        self.output_sorted = self.strSource + 'warning_sorted_filename.csv'
        # 处理included_dirs后的数据（去除重复数据）
        self.strOutput_sorted_include_dir = self.strSource + 'warning_sorted_filename_include_dir.csv'
        # 处理excluded后的数据（去除重复数据）
        self.strOutput_sorted_exclude = self.strSource + 'warning_sorted_filename_exclude.csv'


    def dealWithLog(self):

        included_dirs = ['\\Sword3\\Source\\KG3DEngineDX11',
                         '\\Sword3\\Source\\KG3DEngine\\',
                         '\\Sword3\\Source\\Common\\Schema\\Semantic']  # 包含的目录
        for dir in included_dirs:
            print(dir)
        excluded_dirs = [r'\Sword3\Source\KG3DEngineDX11\KG3DEngineE\Internal\Tool',
                         r'\Sword3\Source\KG3DEngineDX11\KG3DEngineE\Internal\Module\KG3DEngineDX11E\EditTool',
                         r'\Sword3\Source\KG3DEngineDX11\KG3DEngineE\Internal\Module\IPPVideo\detours',
                         r'\Sword3\Source\KG3DEngineDX11\DevEnv',
                         r'\Sword3\Source\KG3DEngineDX11\KG3DEngineE\Internal\Component\KG3D_DebugConsole',
                         r'\Sword3\Source\KG3DEngine\KG3DEngine\JpegLib',
                         r'\Sword3\Source\KG3DEngineDX11\KG3DEngineE\Internal\Module\KG3D_NVBlastDestruction\MeshFix',
                         r'\Sword3\Source\KG3DEngineDX11\KG3DEngineE\Internal\Module\IPPVideo\ffmpeg',
                         r'\Sword3\Source\KG3DEngine\KG3DSound\KMp3LibClass\src\KMp3HufTable.h',
                         r'\Sword3\Source\KG3DEngineDX11\KG3DEngineE\Internal\Module\KG3D_Wwise',
                         r'\Sword3\Source\KG3DEngineDX11\KG3DEngineE\Publish\Include\IKG3D_WwiseInterface.h',
                         r'\Sword3\Source\KG3DEngine\KG3DEngineCommon\kg3d_DataCenter\datacenter\external\boost',
                         r'\Sword3\Source\KG3DEngine\KG3DEngineCommon\kg3d_DataCenter\datacenter\lua',
                         r'\Sword3\Source\KG3DEngine\KG3DEngineCommon\include\KG3D_DataCenter\datacenterpub\external']  # 去掉的目录

        excluded_Cwarning = ['C4189', 'C26475', 'C26814', 'C26440', 'C4100', 'C4251']
        excluded_StrWarning = ["(con.4)",
                               "Use 'nullptr' rather than 0 or NULL (es.47)",
                               "Avoid 'goto' (es.76)", "Avoid 'goto' (es.76)",
                               "Don't use C-style casts (type.4)",
                               "Consider using gsl::finally if final action is intended (gsl.util)",
                               "Default constructor may not throw. Declare it 'noexcept' (f.6)",
                               "Prefer to use gsl::at() instead of unchecked subscript operator (bounds.4)"
                               ]

        excluded_dirs_Cwarning=[
            {'dir':r'\Sword3\Source\KG3DEngineDX11\KG3DEngineE\Internal\Module\KG3D_NVBlastDestruction\src\KG3D_DestructionFamily.h','type':'C26812'}
        ]
        header = ['id', 'file', 'row', 'col', 'warning']  # 定义表格的列名
        with open(self.strOutput_file, 'w', newline='') as f:  # 创建空的 CSV 文件并写入表头
            writer = csv.writer(f)
            writer.writerow(header)

        with open(self.strOutput_file_include_dir, 'w', newline='') as f:  # 创建空的 CSV 文件并写入表头
            writer = csv.writer(f)
            writer.writerow(header)

        with open(self.strOutput_file_exclude, 'w', newline='') as f:  # 创建空的 CSV 文件并写入表头
            writer = csv.writer(f)
            writer.writerow(header)

        data_hold_exclude = []
        data_hold_include_dir = []
        data_hold = []
        with open(self.strInputSrc, 'r') as file:
            for line in file:
                if 'warning' in line:
                    data = {}
                    match = re.search(r'warning\s+(C\d+):', line)  # 找警告类型
                    if match:
                        warning_type = match.group(1)
                        data['id'] = warning_type
                        '''
                        if any(dir in warning_type for dir in excluded_Cwarning): #过滤掉C类警告
                            continue
                        '''
                    else:
                        continue

                    match = re.search(r'[A-Za-z]:\\[\\\w\s\-.]*\.\w+', line)  # 找路径
                    if match:
                        path = match.group()
                        match_ = re.search(r'\\Sword3\\(.*)', path)  # 匹配从sword3开始到最后
                        if match_:
                            path = match_.group()
                            data['file'] = path
                        else:
                            continue
                    else:
                        match1 = re.search(r'[A-Za-z]:\\(?:[^\\/:*?"<>|\r\n]+\\)*([^\\/:*?"<>|\r\n]+)',
                                           line)  # 匹配没有后缀的路径
                        if match1:
                            path = match1.group()
                            match2 = re.search(r'\\Sword3\\(.*)', path)  # 匹配从sword3开始到最后
                            if match2:
                                path = match2.group()
                                data['file'] = path
                            else:
                                continue
                        else:
                            data['file'] = " "
                            print("有空白路径")

                    match = re.search(r'\((\d+),(\d+)\):', line)  # 行和列
                    if match:
                        row = str(match.group(1))
                        col = str(match.group(2))
                        data['row'] = row
                        data['col'] = col
                    else:
                        match_ = re.search(r'\((\d+)\):', line)
                        if match_:
                            row = match_.group(1)
                            data['row'] = row
                            data['col'] = " "
                        else:
                            data['row'] = " "
                            data['col'] = " "
                            print("有空白行列")

                    match = re.search(r'warning\s+([^:]+):\s+(.*)', line)  # 警告内容
                    if match:
                        warning_info = match.group(2)
                        '''
                        if any(str in warning_info for str in excluded_StrWarning):  # 过滤掉excluded_StrWarning里面的警告
                            continue
                        '''
                        data['warning'] = warning_info
                    else:
                        data['warning'] = " "
                        print("有空白警告")

                    '''
                    if any(dir in data_exclude['file'] for dir in included_dirs) and not any(   #这里是有些bug吗？？？？
                            dir in data_exclude['file'] for dir in excluded_dirs):
    '''

                    data_hold.append(data)
                    if any(dir in data['file'] for dir in included_dirs):  # include dir
                        data_hold_include_dir.append(data)
                        if any(dir in data['file'] for dir in excluded_dirs):  # 过滤路径
                            continue

                        if any(dir in warning_type for dir in excluded_Cwarning):  # 过滤掉C类警告
                            continue

                        if any(str in warning_info for str in excluded_StrWarning):  # 过滤掉excluded_StrWarning里面的警告
                            continue

                        #过滤掉excluded_dirs_Cwarning里面的警告 同时成立再过滤
                        bFlag=False
                        for dic_excludedInfo in excluded_dirs_Cwarning:
                            if dic_excludedInfo['dir'] in data['file'] and dic_excludedInfo['type'] in warning_type:
                                bFlag=True
                                break
                        if bFlag:
                            continue

                        data_hold_exclude.append(data)

        with open(self.strOutput_file, 'a', newline='') as f:
            writer = csv.writer(f)
            for row in data_hold:
                writer.writerow([row['id'], row['file'], row['row'], row['col'], row['warning']])  # 写入每一行数据

        with open(self.strOutput_file_include_dir, 'a', newline='') as f:
            writer = csv.writer(f)
            for row in data_hold_include_dir:
                writer.writerow([row['id'], row['file'], row['row'], row['col'], row['warning']])  # 写入每一行数据

        with open(self.strOutput_file_exclude, 'a', newline='') as f:
            writer = csv.writer(f)
            for row in data_hold_exclude:
                writer.writerow([row['id'], row['file'], row['row'], row['col'], row['warning']])  # 写入每一行数据

        print(f"data_hold len:{len(data_hold)}")
        print(f"data_hold_include_dir len:{len(data_hold_include_dir)}")
        print(f"data_hold_exclude len:{len(data_hold_exclude)}")

    def delete_duplicates_and_sortbyfile(self):
        df = pd.read_csv(self.strOutput_file)
        df.drop_duplicates(inplace=True)
        sorted_df = df.sort_values(by='file')
        sorted_df.to_csv(self.output_sorted, index=False)
        print(f"data_hold_duplicates len:{len(df)}")

        df = pd.read_csv(self.strOutput_file_include_dir)
        df.drop_duplicates(inplace=True)
        sorted_df = df.sort_values(by='file')
        sorted_df.to_csv(self.strOutput_sorted_include_dir, index=False)
        print(f"data_hold_include_dir_duplicates len:{len(df)}")

        df = pd.read_csv(self.strOutput_file_exclude)
        df.drop_duplicates(inplace=True)
        sorted_df = df.sort_values(by='file')
        sorted_df.to_csv(self.strOutput_sorted_exclude, index=False)
        print(f"data_hold_exclude_duplicates len:{len(df)}")

    def get_last_modified_name(self,strFilePath, nLineNumber, list_lines, bFlag=False):
        try:
            # 运行 svn blame 命令并捕获输出
            if not bFlag:
                strFilePath = strFilePath.replace('\\', '/')
                strFilePath = strFilePath[7:]
                print(strFilePath)
                strURL = 'https://xsjreposvr1.seasungame.com/svn/Sword3/trunk' + strFilePath
                # list_command = ["svn", "blame",'--use-merge-history',strURL]
                command = f'svn blame --use-merge-history {strURL} -x --ignore-eol-style'
                # command = f'svn blame --use-merge-history {strURL} -x --ignore-eol-style'
                pi = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE)
                res = pi.stdout.read()
                try:
                    res = str(res, encoding='gbk')
                except:
                    res = str(res, encoding='utf8')
                # print(res)
                # 按行分割输出内容
                list_lines = res.splitlines()

            # 获取指定行的信息

            target_line = list_lines[nLineNumber - 1]
            # 提取最后一次修改的人员名称
            list_name = target_line.split(' ')
            '''
            print(list_name)
            for i in range(1,len(list_name)):
                if re.match(r'\d{1,9}',list_name[i]):
                    break
            #print(i)
            for j in range(i+1,len(list_name)):
                if list_name[j]!='':
                    break
            strName = list_name[j]
            '''
            list_name_new = []
            for strName in list_name:
                if strName != '':
                    list_name_new.append(strName)
            # print(list_name)
            strName = list_name_new[1]
            if list_name_new[0] == 'G':
                strName = list_name_new[2]
            print(f"name:{strName}  number:{nLineNumber}  target_line:{target_line}----")
            # print(target_line.split(' '))
            return strName, list_lines
        except (IndexError, subprocess.CalledProcessError) as e:
            print(f"Error: {e}  path: {strFilePath}  number:{nLineNumber}")
            return "代码更新", list_lines

    def dealwithSourceData(self):
        header = ['Id', 'File', 'Count', 'Row', 'Col', 'Lastest modified', 'Warning']  # 定义表格的列名

        df = pd.read_csv(self.strOutput_sorted_exclude)
        from openpyxl import Workbook
        from openpyxl.styles import Font
        # 合并单元格
        from openpyxl.styles import Alignment
        # 创建居中格式的样式对象
        center_alignment = Alignment(horizontal='left', vertical='center')
        # 创建加粗的字体样式对象
        bold_font = Font(bold=True)

        wb = Workbook()
        ws = wb.active

        list_fileInfo = []
        # 将DataFrame数据写入Excel工作表
        for i in range(1, len(header) + 1):
            ws.cell(row=1, column=i).value = header[i - 1]
            ws.cell(row=1, column=i).font = bold_font
        for r, row in enumerate(df.values, start=1):
            for c, value in enumerate(row, start=1):
                if c >= 3:
                    c = c + 1
                if c >= 6:
                    c = c + 1
                ws.cell(row=r + 1, column=c).value = value

        # count列全部置为1
        strPrevFilePath = ""
        list_lines = []
        for i in range(2, ws.max_row + 1):
            ws.cell(row=i, column=3).value = 1
            strFilePath = ws.cell(row=i, column=2).value
            nLineNumber = int(ws.cell(row=i, column=4).value)
            try:
                if strPrevFilePath != strFilePath:
                    strPrevFilePath = strFilePath
                    strName, list_lines = self.get_last_modified_name(strFilePath, nLineNumber, list_lines, False)
                else:
                    strName, list_lines = self.get_last_modified_name(strFilePath, nLineNumber, list_lines, True)

                ws.cell(row=i, column=6).value = strName
            except:
                pass

        fileNameLast = None
        nIndexStart = 0
        nIndexEnd = 0
        nRowCounter = 0
        print(ws, ws.max_row + 1)
        # 处理最后一行 2
        for i in range(2, ws.max_row + 2):
            if not fileNameLast:
                fileNameLast = ws[f"B{i}"].value
                print(fileNameLast)
                nIndexStart = i
                nIndexEnd = i
                # 处理第一行
                list_fileInfo.append({"file": ws[f'B{nIndexStart}'].value, "count": 1})
                continue
            fileName = ws[f"B{i}"].value
            if fileName == fileNameLast:
                nIndexEnd += 1
            else:
                ws.merge_cells(f'B{nIndexStart}:B{nIndexEnd}')
                ws.merge_cells(f'C{nIndexStart}:C{nIndexEnd}')
                print(f'B{nIndexStart}:B{nIndexEnd}')
                # 单元格设置居中
                ws[f'B{nIndexStart}'].alignment = center_alignment
                ws[f'C{nIndexStart}'].alignment = center_alignment
                ws[f'C{nIndexStart}'].value = nIndexEnd - nIndexStart + 1
                list_fileInfo.append({"file": ws[f'B{nIndexStart}'].value, "count": nIndexEnd - nIndexStart + 1})
                # 重置状态
                fileNameLast = fileName
                nIndexStart = i
                nIndexEnd = i
        # 保存Excel文件

        # 设置file列名称长度
        ws.column_dimensions['B'].width = 120
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 150

        # 统计页面
        header = ['file', 'count']
        wsCount = wb.create_sheet(title="fileCount")
        for i in range(1, len(header) + 1):
            wsCount.cell(row=1, column=i).value = header[i - 1]
            wsCount.cell(row=1, column=i).font = bold_font

        # 自动定义排序规则
        def custom_sort_rule(dic_info):
            return dic_info['count']

        list_fileInfo.sort(key=custom_sort_rule, reverse=True)
        print(list_fileInfo)
        for i in range(2, len(list_fileInfo) + 1):
            for j in range(len(header)):
                print(list_fileInfo[i - 2][header[j]])
                wsCount.cell(row=i, column=j + 1).value = list_fileInfo[i - 2][header[j]]
                wsCount.cell(row=i, column=j + 1).alignment = center_alignment

        wsCount.column_dimensions['A'].width = 130
        wb.save(self.strFileName)

    #统计个人warning数量
    def CountLastestModified(self):
        # 读取Excel文件
        self.log.info('CountLastestModified start')
        data = pd.read_excel(self.strFileName, engine='openpyxl')
        # 统计每个姓名的数据条数
        name_counts = data['Lastest modified'].value_counts().sort_values(ascending=False)
        # 创建新的DataFrame
        df = pd.DataFrame({'Lastest modified': name_counts.index, 'Count': name_counts.values})
        # 将新的DataFrame写入到新的Sheet中
        with pd.ExcelWriter(self.strFileName, mode='a', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Modified', index=False)
        self.log.info('CountLastestModified end')


    def CopyDataToServer(self):
        #strServerPath = SERVER_PATH + f"\Warning\{self.strFormattedDate}"
        strServerPath=f"{SERVER_PATH}{os.sep}Warning{os.sep}{self.strFormattedDate}"
        if not filecontrol_existFileOrFolder(strServerPath):
            filecontrol_createFolder(strServerPath)

        #strLocalPath = "E:\AUTO_BVT_NEW\case\liuzhu\py3\TempFolder\warning\\" + self.strFormattedDate
        strLocalPath=f"{os.getcwd()}{os.sep}warning{os.sep}{self.strFormattedDate}"
        self.log.info(strServerPath)
        self.log.info(strLocalPath)
        filecontrol_copyFileOrFolder(strLocalPath, strServerPath)


    def OutPut_data(self):
        strLastData=""
        while True:
            strFormattedDate=datetime.now().strftime("%Y-%m-%d")
            if strLastData==strFormattedDate:
                sleep_heartbeat(30)
            else:
                time_now = time.strftime("%H%M", time.localtime())
                nTime=int(time_now)
                if nTime >= 900:
                    self.SetFilePath(strFormattedDate)
                    while not filecontrol_existFileOrFolder(self.strInputSrc):
                        sleep_heartbeat(10)
                        if strFormattedDate!=datetime.now().strftime("%Y-%m-%d"):
                            strFormattedDate=datetime.now().strftime("%Y-%m-%d")
                            self.SetFilePath(strFormattedDate)
                    self.dealWithLog()
                    self.delete_duplicates_and_sortbyfile()
                    self.dealwithSourceData()
                    self.CountLastestModified()
                    self.CopyDataToServer()
                    #处理数据后重置日期
                    self.log.info(strFormattedDate)
                    strLastData=strFormattedDate
                else:
                    sleep_heartbeat(30)

    def check_dic_args(self, dic_args):
        pass


    def run_local(self, dic_args):
        self.check_dic_args(dic_args)
        self.OutPut_data()


if __name__ == '__main__':
    oob = CaseEngineWarning()
    oob.run_from_IQB()
